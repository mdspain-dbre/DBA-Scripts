"""Generate an Excel workbook documenting the OpenTelemetry sqlserverreceiver
metrics, events, resource attributes, and feature gates.

Source:
  https://github.com/open-telemetry/opentelemetry-collector-contrib/blob/main/receiver/sqlserverreceiver/documentation.md
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT = "sqlserverreceiver_metrics.xlsx"

# --- Data --------------------------------------------------------------------

# (name, description, unit, instrument_type, value_type, monotonic_or_aggregation, default_enabled, stability, windows_only, direct_connect_only, attributes)
DEFAULT_METRICS = [
    ("sqlserver.batch.request.rate",            "Number of batch requests received by SQL Server.",                         "{requests}/s",      "Gauge", "Double", "",           True,  "Development", False, False, ""),
    ("sqlserver.batch.sql_compilation.rate",    "Number of SQL compilations needed.",                                       "{compilations}/s",  "Gauge", "Double", "",           True,  "Development", False, False, ""),
    ("sqlserver.batch.sql_recompilation.rate",  "Number of SQL recompilations needed.",                                     "{compilations}/s",  "Gauge", "Double", "",           True,  "Development", False, False, ""),
    ("sqlserver.lock.wait.rate",                "Number of lock requests resulting in a wait.",                             "{requests}/s",      "Gauge", "Double", "",           True,  "Development", False, False, ""),
    ("sqlserver.lock.wait_time.avg",            "Average wait time for all lock requests that had to wait.",                "ms",                "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.page.buffer_cache.hit_ratio",   "Pages found in the buffer pool without having to read from disk.",         "%",                 "Gauge", "Double", "",           True,  "Development", False, False, ""),
    ("sqlserver.page.checkpoint.flush.rate",    "Number of pages flushed by operations requiring dirty pages to be flushed.","{pages}/s",        "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.page.lazy_write.rate",          "Number of lazy writes moving dirty pages to disk.",                        "{writes}/s",        "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.page.life_expectancy",          "Time a page will stay in the buffer pool.",                                "s",                 "Gauge", "Int",    "",           True,  "Development", False, False, "performance_counter.object_name"),
    ("sqlserver.page.operation.rate",           "Number of physical database page operations issued.",                      "{operations}/s",    "Gauge", "Double", "",           True,  "Development", True,  False, "type (read|write)"),
    ("sqlserver.page.split.rate",               "Number of pages split as a result of overflowing index pages.",            "{pages}/s",         "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.transaction.rate",              "Number of transactions started for the database (not including XTP-only transactions).", "{transactions}/s", "Gauge", "Double", "", True,  "Development", True,  False, ""),
    ("sqlserver.transaction.write.rate",        "Number of transactions that wrote to the database and committed.",         "{transactions}/s",  "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.transaction_log.flush.data.rate","Total number of log bytes flushed.",                                      "By/s",              "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.transaction_log.flush.rate",    "Number of log flushes.",                                                   "{flushes}/s",       "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.transaction_log.flush.wait.rate","Number of commits waiting for a transaction log flush.",                  "{commits}/s",       "Gauge", "Double", "",           True,  "Development", True,  False, ""),
    ("sqlserver.transaction_log.growth.count",  "Total number of transaction log expansions for a database.",               "{growths}",         "Sum",   "Int",    "Cumulative; monotonic=true", True, "Development", True,  False, ""),
    ("sqlserver.transaction_log.shrink.count",  "Total number of transaction log shrinks for a database.",                  "{shrinks}",         "Sum",   "Int",    "Cumulative; monotonic=true", True, "Development", True,  False, ""),
    ("sqlserver.transaction_log.usage",         "Percent of transaction log space used.",                                   "%",                 "Gauge", "Int",    "",           True,  "Development", True,  False, ""),
    ("sqlserver.user.connection.count",         "Number of users connected to the SQL Server.",                             "{connections}",     "Gauge", "Int",    "",           True,  "Development", False, False, ""),
]

OPTIONAL_METRICS = [
    ("sqlserver.computer.uptime",                       "Computer uptime.",                                                  "{seconds}",            "Gauge", "Int",    "",           False, "Development", False, False, ""),
    ("sqlserver.cpu.count",                             "Number of CPUs.",                                                   "{CPUs}",               "Gauge", "Int",    "",           False, "Development", False, False, ""),
    ("sqlserver.database.backup_or_restore.rate",       "Total number of backups/restores.",                                 "{backups_or_restores}/s","Gauge","Double","",           False, "Development", False, False, ""),
    ("sqlserver.database.count",                        "The number of databases.",                                          "{databases}",          "Gauge", "Int",    "",           False, "Development", False, True,  "database.status (online|restoring|recovering|pending_recovery|suspect|offline)"),
    ("sqlserver.database.execution.errors",             "Number of execution errors.",                                       "{errors}",             "Gauge", "Int",    "",           False, "Development", False, False, ""),
    ("sqlserver.database.full_scan.rate",               "The number of unrestricted full table or index scans.",             "{scans}/s",            "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.database.io",                           "The number of bytes of I/O on this file.",                          "By",                   "Sum",   "Int",    "Cumulative; monotonic=true", False,"Development", False, True,  "physical_filename, logical_filename, file_type, direction (read|write)"),
    ("sqlserver.database.latency",                      "Total time that the users waited for I/O issued on this file.",     "s",                    "Sum",   "Double", "Cumulative; monotonic=true", False,"Development", False, True,  "physical_filename, logical_filename, file_type, direction (read|write)"),
    ("sqlserver.database.operations",                   "The number of operations issued on the file.",                      "{operations}",         "Sum",   "Int",    "Cumulative; monotonic=true", False,"Development", False, True,  "physical_filename, logical_filename, file_type, direction (read|write)"),
    ("sqlserver.database.tempdb.space",                 "Total free space in temporary DB.",                                 "KB",                   "Sum",   "Int",    "Cumulative; monotonic=false",False,"Development", False, False, "tempdb.state (free|used)"),
    ("sqlserver.database.tempdb.version_store.size",    "TempDB version store size.",                                        "KB",                   "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.deadlock.rate",                         "Total number of deadlocks.",                                        "{deadlocks}/s",        "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.index.search.rate",                     "Total number of index searches.",                                   "{searches}/s",         "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.lock.timeout.rate",                     "Total number of lock timeouts.",                                    "{timeouts}/s",         "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.lock.wait.count",                       "Cumulative count of lock waits that occurred.",                     "{wait}",               "Sum",   "Int",    "Cumulative; monotonic=true", False,"Development", False, True,  ""),
    ("sqlserver.login.rate",                            "Total number of logins.",                                           "{logins}/s",           "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.logout.rate",                           "Total number of logouts.",                                          "{logouts}/s",          "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.memory.grants.pending.count",           "Total number of memory grants pending.",                            "{grants}",             "Sum",   "Int",    "Cumulative; monotonic=false",False,"Development", False, False, ""),
    ("sqlserver.memory.usage",                          "Total memory in use.",                                              "KB",                   "Sum",   "Double", "Cumulative; monotonic=false",False,"Development", False, False, ""),
    ("sqlserver.os.wait.duration",                      "Total wait time for this wait type.",                               "s",                    "Sum",   "Double", "Cumulative; monotonic=true", False,"Development", False, True,  "wait.category, wait.type"),
    ("sqlserver.page.buffer_cache.free_list.stalls.rate","Number of free list stalls.",                                      "{stalls}/s",           "Gauge", "Int",    "",           False, "Development", False, False, ""),
    ("sqlserver.page.lookup.rate",                      "Total number of page lookups.",                                     "{lookups}/s",          "Gauge", "Double", "",           False, "Development", False, False, ""),
    ("sqlserver.processes.blocked",                     "The number of processes that are currently blocked.",               "{processes}",          "Gauge", "Int",    "",           False, "Development", False, True,  ""),
    ("sqlserver.replica.data.rate",                     "Throughput rate of replica data.",                                  "By/s",                 "Gauge", "Double", "",           False, "Development", False, False, "replica.direction (transmit|receive)"),
    ("sqlserver.resource_pool.disk.operations",         "The rate of operations issued.",                                    "{operations}/s",       "Gauge", "Double", "",           False, "Development", False, True,  "direction (read|write)"),
    ("sqlserver.resource_pool.disk.throttled.read.rate","The number of read operations that were throttled in the last second.","{reads}/s",         "Gauge", "Int",    "",           False, "Development", False, True,  ""),
    ("sqlserver.resource_pool.disk.throttled.write.rate","The number of write operations that were throttled in the last second.","{writes}/s",      "Gauge", "Double", "",           False, "Development", False, True,  ""),
    ("sqlserver.table.count",                           "The number of tables.",                                             "{tables}",             "Sum",   "Int",    "Cumulative; monotonic=false",False,"Development", False, False, "table.state (active|inactive), table.status (temporary|permanent)"),
    ("sqlserver.transaction.delay",                     "Time consumed in transaction delays.",                              "ms",                   "Sum",   "Double", "Cumulative; monotonic=false",False,"Development", False, False, ""),
    ("sqlserver.transaction.mirror_write.rate",         "Total number of mirror write transactions.",                        "{transactions}/s",     "Gauge", "Double", "",           False, "Development", False, False, ""),
]

RESOURCE_ATTRIBUTES = [
    # (name, description, type, default_enabled)
    ("host.name",                "The host name of SQL Server.",                                                                           "Any Str", True),
    ("server.address",           "Name of the database host.",                                                                             "Any Str", False),
    ("server.port",              "Server port number.",                                                                                    "Any Int", False),
    ("service.instance.id",      "A unique identifier of the SQL Server instance in the format host:port. Only when directly connected.",  "Any Str", True),
    ("sqlserver.computer.name",  "The name of the SQL Server instance being monitored.",                                                   "Any Str", False),
    ("sqlserver.database.name",  "The name of the SQL Server database.",                                                                   "Any Str", True),
    ("sqlserver.instance.name",  "The name of the SQL Server instance being monitored.",                                                   "Any Str", False),
]

EVENTS = [
    # (event_name, default_enabled, description)
    ("db.server.query_sample", False, "Per-query sample event capturing currently running / recently sampled queries."),
    ("db.server.top_query",    False, "Aggregated 'top query' event with cumulative execution stats per plan."),
]

EVENT_ATTRIBUTES = [
    # (event_name, attribute, description, type)
    # db.server.query_sample
    ("db.server.query_sample", "client.address",                       "Hostname or address of the client.",                                                              "Any Str"),
    ("db.server.query_sample", "client.port",                          "TCP port used by the client.",                                                                    "Any Int"),
    ("db.server.query_sample", "db.namespace",                         "The database name.",                                                                              "Any Str"),
    ("db.server.query_sample", "db.query.text",                        "The text of the database query being executed.",                                                  "Any Str"),
    ("db.server.query_sample", "db.system.name",                       "The DBMS product as identified by the client instrumentation.",                                   "Any Str"),
    ("db.server.query_sample", "network.peer.address",                 "IP address of the peer client.",                                                                  "Any Str"),
    ("db.server.query_sample", "network.peer.port",                    "TCP port used by the peer client.",                                                               "Any Int"),
    ("db.server.query_sample", "sqlserver.blocking_session_id",        "Session ID that is blocking the current session. 0 if none.",                                     "Any Int"),
    ("db.server.query_sample", "sqlserver.blocking.start_time",        "Timestamp of when the current blocking wait began (ISO 8601).",                                   "Any Str"),
    ("db.server.query_sample", "sqlserver.context_info",               "Context information for the session, hex string.",                                                "Any Str"),
    ("db.server.query_sample", "sqlserver.command",                    "SQL command type being executed.",                                                                "Any Str"),
    ("db.server.query_sample", "sqlserver.cpu_time",                   "CPU time consumed by the query, in seconds.",                                                     "Any Double"),
    ("db.server.query_sample", "sqlserver.deadlock_priority",          "Deadlock priority value for the session.",                                                        "Any Int"),
    ("db.server.query_sample", "sqlserver.estimated_completion_time",  "Estimated time remaining for the request to complete, in seconds.",                               "Any Double"),
    ("db.server.query_sample", "sqlserver.lock_timeout",               "Lock timeout value in seconds.",                                                                  "Any Double"),
    ("db.server.query_sample", "sqlserver.logical_reads",              "Number of logical reads (data read from cache/memory).",                                          "Any Int"),
    ("db.server.query_sample", "sqlserver.open_transaction_count",     "Number of transactions currently open in the session.",                                           "Any Int"),
    ("db.server.query_sample", "sqlserver.percent_complete",           "Percentage of work completed.",                                                                   "Any Double"),
    ("db.server.query_sample", "sqlserver.query_hash",                 "Hash of the query text (HEX).",                                                                   "Any Str"),
    ("db.server.query_sample", "sqlserver.query_plan_hash",            "Hash of the query plan (HEX).",                                                                   "Any Str"),
    ("db.server.query_sample", "sqlserver.query_start",                "Timestamp of when the SQL query started (ISO 8601).",                                             "Any Str"),
    ("db.server.query_sample", "sqlserver.reads",                      "Number of physical reads performed by the query.",                                                "Any Int"),
    ("db.server.query_sample", "sqlserver.request_status",             "Status of the request (e.g., running, suspended).",                                               "Any Str"),
    ("db.server.query_sample", "sqlserver.wait.resource.id",           "SQL Server identifier for the locked or waited-on resource, if available.",                       "Any Str"),
    ("db.server.query_sample", "sqlserver.wait.resource.type",         "SQL Server type of the locked or waited-on resource, if available.",                              "Any Str"),
    ("db.server.query_sample", "sqlserver.row_count",                  "Number of rows affected or returned by the query.",                                               "Any Int"),
    ("db.server.query_sample", "sqlserver.session_id",                 "ID of the SQL Server session.",                                                                   "Any Int"),
    ("db.server.query_sample", "sqlserver.session_status",             "Status of the session (e.g., running, sleeping).",                                                "Any Str"),
    ("db.server.query_sample", "sqlserver.total_elapsed_time",         "Total elapsed time for completed executions of this plan, delta seconds.",                        "Any Double"),
    ("db.server.query_sample", "sqlserver.transaction_id",             "Unique ID of the active transaction.",                                                            "Any Int"),
    ("db.server.query_sample", "sqlserver.transaction_isolation_level","Transaction isolation level used in the session (numeric constant).",                             "Any Int"),
    ("db.server.query_sample", "sqlserver.wait_resource",              "The resource for which the session is waiting.",                                                  "Any Str"),
    ("db.server.query_sample", "sqlserver.wait_time",                  "Duration in seconds the request has been waiting.",                                               "Any Double"),
    ("db.server.query_sample", "sqlserver.wait_type",                  "Type of wait encountered by the request. Empty if none.",                                         "Any Str"),
    ("db.server.query_sample", "sqlserver.writes",                     "Number of writes performed by the query.",                                                        "Any Int"),
    ("db.server.query_sample", "user.name",                            "Login name associated with the SQL Server session.",                                              "Any Str"),
    ("db.server.query_sample", "sqlserver.procedure_id",               "The SQL Server ID of the stored procedure, if any.",                                              "Any Str"),
    ("db.server.query_sample", "sqlserver.procedure_name",             "The name of the stored procedure, if any.",                                                      "Any Str"),
    # db.server.top_query
    ("db.server.top_query",    "sqlserver.total_worker_time",          "Total CPU time consumed by executions of this plan, delta seconds.",                              "Any Double"),
    ("db.server.top_query",    "db.query.text",                        "The text of the database query being executed.",                                                  "Any Str"),
    ("db.server.top_query",    "sqlserver.execution_count",            "Number of times the plan has executed since last compiled, delta value.",                         "Any Int"),
    ("db.server.top_query",    "sqlserver.total_logical_reads",        "Total logical reads by executions of this plan since compiled, delta value.",                     "Any Int"),
    ("db.server.top_query",    "sqlserver.total_logical_writes",       "Total logical writes by executions of this plan since compiled, delta value.",                    "Any Int"),
    ("db.server.top_query",    "sqlserver.total_physical_reads",       "Total physical reads by executions of this plan since compiled, delta value.",                    "Any Int"),
    ("db.server.top_query",    "sqlserver.query_hash",                 "Hash of the query text (HEX).",                                                                   "Any Str"),
    ("db.server.top_query",    "sqlserver.query_plan",                 "The query execution plan used by SQL Server.",                                                    "Any Str"),
    ("db.server.top_query",    "sqlserver.query_plan_hash",            "Hash of the query execution plan (HEX).",                                                         "Any Str"),
    ("db.server.top_query",    "sqlserver.total_rows",                 "Total rows returned by the query, delta value.",                                                  "Any Int"),
    ("db.server.top_query",    "sqlserver.total_elapsed_time",         "Total elapsed time for completed executions of this plan, delta seconds.",                        "Any Double"),
    ("db.server.top_query",    "sqlserver.total_grant_kb",             "Total reserved memory grant in KB this plan received since compiled, delta value.",               "Any Int"),
    ("db.server.top_query",    "server.address",                       "Network address of the server hosting the database.",                                             "Any Str"),
    ("db.server.top_query",    "server.port",                          "Port number on which the server is listening.",                                                   "Any Int"),
    ("db.server.top_query",    "db.system.name",                       "DBMS product as identified by the client instrumentation.",                                       "Any Str"),
    ("db.server.top_query",    "sqlserver.procedure_execution_count",  "Number of times the procedure executed since last compiled, delta value.",                        "Any Int"),
    ("db.server.top_query",    "sqlserver.procedure_id",               "The SQL Server ID of the stored procedure, if any.",                                              "Any Str"),
    ("db.server.top_query",    "sqlserver.procedure_name",             "The name of the stored procedure, if any.",                                                      "Any Str"),
]

FEATURE_GATES = [
    # (id, lifecycle_stage, description, from_version, to_version)
    ("receiver.sqlserver.RemoveServerResourceAttribute", "alpha",
     "When enabled, the server.address and server.port resource attributes are removed from metrics.",
     "v0.129.0", "N/A"),
]


# --- Workbook construction ---------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
DEFAULT_TRUE_FILL = PatternFill("solid", fgColor="E2EFDA")
DEFAULT_FALSE_FILL = PatternFill("solid", fgColor="FFF2CC")


def write_sheet(ws, headers, rows, default_enabled_col_idx=None):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        ws.append(list(row))
    # Highlight default-enabled column
    if default_enabled_col_idx is not None:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=default_enabled_col_idx)
            cell.fill = DEFAULT_TRUE_FILL if cell.value is True else DEFAULT_FALSE_FILL
    # Wrap text + autosize-ish
    for col_idx, _ in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(14, max_len + 2), 60)
        for r in range(1, ws.max_row + 1):
            ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def main():
    wb = Workbook()

    # README sheet
    ws = wb.active
    ws.title = "README"
    readme = [
        ["OpenTelemetry sqlserverreceiver — metrics & events reference"],
        [],
        ["Source",
         "https://github.com/open-telemetry/opentelemetry-collector-contrib/blob/main/receiver/sqlserverreceiver/documentation.md"],
        ["Generated", "Snapshot of receiver documentation."],
        [],
        ["Sheet", "Contents"],
        ["Default Metrics",  f"{len(DEFAULT_METRICS)} metrics emitted by default. Disable per-metric in collector config."],
        ["Optional Metrics", f"{len(OPTIONAL_METRICS)} metrics NOT emitted by default. Enable per-metric in collector config."],
        ["Resource Attrs",   f"{len(RESOURCE_ATTRIBUTES)} resource attributes attached to all telemetry from this receiver."],
        ["Events",           f"{len(EVENTS)} log/event types (currently both optional)."],
        ["Event Attributes", f"{len(EVENT_ATTRIBUTES)} attributes across all events."],
        ["Feature Gates",    f"{len(FEATURE_GATES)} feature gates."],
        [],
        ["Notes",
         "windows_only=TRUE means metric only collected on Windows host. "
         "direct_connect_only=TRUE means receiver must be configured to connect directly to SQL Server (vs. agent mode)."],
        [],
        ["Enable a metric",  "metrics:\n  <metric_name>:\n    enabled: true"],
        ["Disable a metric", "metrics:\n  <metric_name>:\n    enabled: false"],
    ]
    for row in readme:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 110
    for r in range(1, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    ws["A1"].font = Font(bold=True, size=14)

    # Metrics columns
    metric_headers = [
        "Metric Name", "Description", "Unit", "Instrument", "Value Type",
        "Aggregation / Monotonicity", "Default Enabled", "Stability",
        "Windows Only", "Direct-Connect Only", "Attributes",
    ]

    ws_def = wb.create_sheet("Default Metrics")
    write_sheet(ws_def, metric_headers, DEFAULT_METRICS, default_enabled_col_idx=7)

    ws_opt = wb.create_sheet("Optional Metrics")
    write_sheet(ws_opt, metric_headers, OPTIONAL_METRICS, default_enabled_col_idx=7)

    # Resource attributes
    ws_ra = wb.create_sheet("Resource Attrs")
    write_sheet(
        ws_ra,
        ["Attribute", "Description", "Type", "Default Enabled"],
        RESOURCE_ATTRIBUTES,
        default_enabled_col_idx=4,
    )

    # Events summary
    ws_ev = wb.create_sheet("Events")
    write_sheet(
        ws_ev,
        ["Event Name", "Default Enabled", "Description"],
        EVENTS,
        default_enabled_col_idx=2,
    )

    # Event attributes
    ws_ea = wb.create_sheet("Event Attributes")
    write_sheet(
        ws_ea,
        ["Event", "Attribute", "Description", "Type"],
        EVENT_ATTRIBUTES,
    )

    # Feature gates
    ws_fg = wb.create_sheet("Feature Gates")
    write_sheet(
        ws_fg,
        ["Gate ID", "Stage", "Description", "From Version", "To Version"],
        FEATURE_GATES,
    )

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"  Default metrics:   {len(DEFAULT_METRICS)}")
    print(f"  Optional metrics:  {len(OPTIONAL_METRICS)}")
    print(f"  Resource attrs:    {len(RESOURCE_ATTRIBUTES)}")
    print(f"  Events:            {len(EVENTS)}")
    print(f"  Event attributes:  {len(EVENT_ATTRIBUTES)}")
    print(f"  Feature gates:     {len(FEATURE_GATES)}")


if __name__ == "__main__":
    main()
